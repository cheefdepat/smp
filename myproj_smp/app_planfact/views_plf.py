from django.shortcuts import render, get_object_or_404, redirect
from .models import PlanFactTab
# import datetime
from django.core.paginator import Paginator

import logging

import openpyxl
from io import BytesIO
from django.http import HttpResponse
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side

# ------------------------------------------------- Зона логгирования ----------------
class MyFilter(logging.Filter):
    def filter(self, record):
        return not record.getMessage().startswith('GET') and not record.getMessage().startswith('POST')

# Создайте логгер
logger = logging.getLogger(__name__)

# Установите уровень логирования
logger.setLevel(logging.INFO)

# Создайте обработчик логирования для файла
file_handler = logging.FileHandler('log_plf.txt')
file_handler.addFilter(MyFilter())

# Добавьте обработчик логирования к логгеру
logger.addHandler(file_handler)

# -------------------------------------------------  логгирования -----------------------

def zamena_pustot(pole_proverki_daty):
    # Функция проверки является ли значение в поле ДАТЫ - пустым??? для ВСЕХ ДАТ проверку!!!!
    # Проверяем ДАТУ, является ли значение None ---------------- для ВСЕХ ДАТ проверку!!!!
    if pole_proverki_daty == "":
        pole_proverki_daty = None
    else:
        pole_proverki_daty = pole_proverki_daty
    return pole_proverki_daty


def plf_start_list(request):
    # --------------cd
    user_groups_list = []
    for i in request.user.groups.all():
        print(i)
        user_groups_list.append(str(i))
    # groups = Group.objects.all()
    print(user_groups_list)

    # --------------

    is_vps_group = request.user.groups.filter(name='vps').exists()
    query_fio = request.GET.get('search_fio', '')  # Получаем строку поиска по FIO
    query_kurir = request.GET.get('search_kurir', '')  # Получаем строку поиска по курирующему подразделению
    query_otrabot = request.GET.get('search_otrabot', '')  # Получаем строку поиска по отработанным записям
    query_data_zamecaniya = request.GET.get('search_data_zamecaniya', '')  # Получаем строку поиска по отработанным записям
    records_per_page = request.GET.get('records_per_page', 20)  # Получаем количество записей на странице
    print(query_data_zamecaniya)
    # Фильтруем данные по обоим полям и сортируем по p_p

    print(request.user.groups)

    # if user_groups_list == ['vps']:   # ------------------------- ВКЛЮЧИТЬ визмость ОТПРАВКИ В КЭР
    if 'vps' in user_groups_list and len(user_groups_list) == 1:  # ------------------------- ВКЛЮЧИТЬ визмость ОТПРАВКИ В КЭР
    # if user_groups_list == ['#']:     # ------------------------- отключить визмость ОТПРАВКИ В КЭР
        # Исключаем записи, у которых ok_vps равно "ВПС"
        data_plf = PlanFactTab.objects.filter(ok_status_zapolnenia="Заполняется ВПС").order_by('data_planfakta',
                                                                                  'fio_pacienta')  # Сортировка по возрастанию
    else:
        data_plf = PlanFactTab.objects.all().order_by('data_planfakta', 'fio_pacienta')  # Сортировка по возрастанию
    total_records = data_plf.count()  # Общее количество записей

    if query_fio:
        data_plf = data_plf.filter(fio_pacienta__icontains=query_fio)

    if query_kurir:
        data_plf = data_plf.filter(ovpp_name__icontains=query_kurir)

    if query_otrabot:
        data_plf = data_plf.filter(ok_status_zapolnenia=query_otrabot)

    if query_data_zamecaniya:
        data_plf = data_plf.filter(data_planfakta=query_data_zamecaniya)


    # data_smp = data_smp.filter(ok_vps=query_kurir)
    # Получаем уникальные значения для выпадающего списка
    unique_kurir = PlanFactTab.objects.values_list('ovpp_name', flat=True).distinct()
    unique_otrab = PlanFactTab.objects.values_list('ok_status_zapolnenia', flat=True).distinct()

    paginator = Paginator(data_plf, records_per_page)  # Показывать 10 записей на странице
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    filter_records = data_plf.count()
    return render(request, 'planfact_home_page.html', {
        'data_plf': page_obj,
        'paginator':paginator,
        'search_fio': query_fio,
        'search_kurir': query_kurir,
        'search_otrabot': query_otrabot,
        'records_per_page': records_per_page,
        'total_records': total_records,  # Передаем общее количество записей
        'unique_kurir': unique_kurir,  # Передаем уникальные значения в контекст по курир. филиалу ВПС
        'unique_otrab': unique_otrab,  # Передаем уникальные значения в контекст по отработанным в КЭР
         'search_data_zamecaniya': query_data_zamecaniya,  # Передаем уникальные значения в контекст по отработанным в КЭР

        'groups': user_groups_list, # Получаем все группы
        'filter_records': filter_records,  ## количество записей - отфильтрованное
    })


def plf_edit_patient(request, id):
    # -------------- группу передадим ---
    user_groups_list = []
    for i in request.user.groups.all():
        user_groups_list.append(str(i))
    # --------------

    patient = get_object_or_404(PlanFactTab, id=id)

    # --------------- вычсление перес=менных по ДНЯМ!!! ------------
    if request.method == 'POST':

        # ---------   Заполняем поля, которые НУЖНО считать со страницы и сохранять ------------
        logger.info(f'пользователь {request.user} вошел к пациенту {patient.fio_pacienta} -- {patient.data_planfakta}')



        # ------------------- предотвращение ошибки из-за пустоты в ДАТЕ ----
        patient.data_vk = request.POST.get('data_vk')
        patient.data_vk = zamena_pustot(patient.data_vk)

        patient.name_mo_provod_vk = request.POST.get('name_mo_provod_vk')
        patient.diagnoz_mkb10 = request.POST.get('diagnoz_mkb10')
        patient.komment_k_pervichnomu_vizitu = request.POST.get('komment_k_pervichnomu_vizitu')
        patient.komment_k_posled_vizitu = request.POST.get('komment_k_posled_vizitu')
        patient.tyagostnaya_simptomatika_pall_potrebnosti = request.POST.get('tyagostnaya_simptomatika_pall_potrebnosti')
        patient.plan_nablyudeniya_sootvestvuet_tyazhesti_sostoyaniya_i_prognozu = request.POST.get('plan_nablyudeniya_sootvestvuet_tyazhesti_sostoyaniya_i_prognozu')


        patient.vizov_smp_posle_last_vizit = request.POST.get('vizov_smp_posle_last_vizit')

        patient.nalichie_pokazanij_k_okazaniyu_specPMP = request.POST.get('nalichie_pokazanij_k_okazaniyu_specPMP')
        patient.nablyudenie_v_drugoj_mo_parallel_s_ovpp = request.POST.get('nablyudenie_v_drugoj_mo_parallel_s_ovpp')
        patient.name_drugoj_mo_parallel_s_ovpp = request.POST.get('name_drugoj_mo_parallel_s_ovpp')
        patient.name_drugoj_mo_parallel_s_ovpp = request.POST.get('name_drugoj_mo_parallel_s_ovpp')

        patient.osnovaniya_parallel_nabluden = request.POST.get('osnovaniya_parallel_nabluden')

        patient.vypiska_receptov_vps = request.POST.get('vypiska_receptov_vps')
        patient.vypiska_receptov_drugoj_mo = request.POST.get('vypiska_receptov_drugoj_mo')
        patient.potrebnost_v_respiratorke = request.POST.get('potrebnost_v_respiratorke')
        patient.vyvlennye_defekty = request.POST.get('vyvlennye_defekty')
        patient.mery_prinyatye_vps_dlya_uluchsheniya_pmp = request.POST.get('mery_prinyatye_vps_dlya_uluchsheniya_pmp')
        patient.predlozheniya_dlya_uluchsheniya_pmp = request.POST.get('predlozheniya_dlya_uluchsheniya_pmp')
        patient.diagnoz_mkb10 = request.POST.get('diagnoz_mkb10')
        # patient.komment_zav_vps = request.POST.get('komment_zav_vps')
        # patient.komment_aup = request.POST.get('komment_aup')

        patient.otvet_kc = request.POST.get('otvet_kc')


        # ------------------- предотвращение ошибки из-за пустоты в ДАТЕ ----
        patient.data_vklyucheniya_v_reestr = request.POST.get('data_vklyucheniya_v_reestr')
        patient.data_vklyucheniya_v_reestr = zamena_pustot(patient.data_vklyucheniya_v_reestr)
        # ------------------- предотвращение ошибки из-за пустоты в ДАТЕ ----
        patient.data_vklyucheniya_v_reestr = request.POST.get('data_vklyucheniya_v_reestr')
        patient.data_vklyucheniya_v_reestr = zamena_pustot(patient.data_vklyucheniya_v_reestr)
        # ------------------- предотвращение ошибки из-за пустоты в ДАТЕ ----
        patient.data_pervichnogo_vizita = request.POST.get('data_pervichnogo_vizita')
        patient.data_pervichnogo_vizita = zamena_pustot(patient.data_pervichnogo_vizita)
# ------------------- предотвращение ошибки из-за пустоты в ДАТЕ ----
        patient.data_poslednego_vizita = request.POST.get('data_poslednego_vizita')
        patient.data_poslednego_vizita = zamena_pustot(patient.data_poslednego_vizita)

# ------------------- предотвращение ошибки из-за пустоты в ДАТЕ ----
        patient.data_poslednego_poseshcheniya_drugoj_mo = request.POST.get('data_poslednego_poseshcheniya_drugoj_mo')
        patient.data_poslednego_poseshcheniya_drugoj_mo = zamena_pustot(patient.data_poslednego_poseshcheniya_drugoj_mo)






        if 'plf_save_vps' in request.POST:  # Кнопка "Сохранить"
            print('109-----------------')
            logger.info(
                f'пользователь {request.user} сохранил инф-ю о {patient.fio_pacienta} -- {patient.data_planfakta}')
            patient.save()
            return redirect('app_planfact:plf_start_list')

        elif 'ot_ker_nazad_v_vps' in request.POST:  # Кнопка "назад на ВПС"
            patient.ok_status_zapolnenia = "Заполняется ВПС"
            patient.save()
            return redirect('app_planfact:plf_start_list')  # Возвращаем на страницу редактирования


        elif 'plf_send_to_ker' in request.POST:  # Кнопка "Отправить в КЭР"
            # patient.ok_status_zapolnenia = "Передано в КЭР"

            patient.save()  # Сохраняем изменения
            return redirect('app_planfact:plf_proverka_to_ker', id=patient.id)  # Переходим на страницу проверки


    return render(request, 'plf_edit_pac_short.html', {'patient': patient,
                                                       'groups': user_groups_list,  # Получаем все группы

                                                       })
def plf_proverka_to_ker(request, id):
    patient = get_object_or_404(PlanFactTab, id=id)

    if request.method == 'POST':
        if 'ot_vps_v_ker' in request.POST:  # Кнопка "Отправить в КЭР"
            patient.ok_status_zapolnenia = "Передано в КЭР"
            logger.info(
                f'пользователь {request.user} отправил запись о {patient.fio_pacienta} -- {patient.data_planfakta} в КЭР')
            patient.save()
            return redirect('app_planfact:plf_start_list')  # Переходим на главную страницу

        elif 'korrektirovat' in request.POST:  # Кнопка "Корректировать"
            return redirect('app_planfact:plf_edit_patient', id=id)  # Возвращаем на страницу редактирования

    return render(request, 'plf_patient_detail.html', {'patient': patient,
                                                       'group':'ker1'


                                                       })


def plf_export_to_excel(request):
    # Получите данные с учетом фильтров
    search_data_zamecaniya = request.GET.get('search_data_zamecaniya', '')

    # Примените фильтры к вашему запросу
    data_plf = PlanFactTab.objects.all()  # Замените на ваш запрос
    if search_data_zamecaniya:
        data_plf = data_plf.filter(data_planfakta=search_data_zamecaniya)  # Замените на ваше поле

    # Создание Excel файла
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Patients Data'

    # Заголовки столбцов
    headers = [
        'Дата план-факта',
        'ID', 'ОМС',    'ФИО',    'ДР',    'ОВПП',
        'Плановая дата визита согласно ЕМИАС/реестру',
        'Фактическая дата визита согласно ЕМИАС',
        'Вопрос для разбора',
        'Дата ВК (ОК)',
        'Наименование МО, проводившей ВК(ОК)',
        'Диагноз МКБ-10',
        'Дата внесения в Реестр',
        'Дата первичного визита',
        'Комментарий к первичному визиту(при необходимости)',
        'Дата последнего визита',
        'Комментарий к последнему визиту(при необходимости)',
        'Имеющаяся тягостная симптоматика, паллиативные потребности',
        'Индивидуальный план наблюдения соотвествует тяжести состояния и прогнозу',
        'Вызовы СМП после последнего визита(даты)',
        'Наличие показаний к оказанию специализированной ПМП',
        'Наблюдение в других МО (параллельно с ОВППМП)',
        'Наименование другой МО, в которой наблюдается пациент параллельно с параллельно с ОВППМП',
        'Дата последнего посещения другой МО',
        'Обоснование параллельного наблюдения иными МО',
        'Выписка рецептов для лечения тягостной симптматики ОВППМП дата последней выписки, наименование препарата, на какой срок или количество препарата',
        'Выписка рецептов для лечения тягостной симптматикидругой МО дата, наименование препарата, на какой срок или количество препарата',
        'Потребность в респираторной поддержке да / нет, при наличии потребности - комментарии, в том числе дата выявления потребности и дата обеспечения респираторным оборудованием',
        'Дефекты, выявленные в процессе разбора',
        'Меры, принятые в ОВППМП для оказания качественной ПМП',
        'Предложения по организационным решениям для улучшения качества оказания ПМП',
        'Коментарий ЦПП',
    ]
    worksheet.append(headers)
    # Установка ширины для 2-го и 15-го столбца
    worksheet.column_dimensions['A'].width = 10  # Установите нужную ширину для 2-го столбца
    worksheet.column_dimensions['B'].width = 25  # Установите нужную ширину для 2-го столбца
    worksheet.column_dimensions['C'].width = 41  # Установите нужную ширину для 2-го столбца
    worksheet.column_dimensions['D'].width = 16  # Установите нужную ширину столбца
    worksheet.column_dimensions['E'].width = 39  # Установите нужную ширину  столбца
    worksheet.column_dimensions['F'].width = 25  # Установите нужную ширину  столбца
    worksheet.column_dimensions['G'].width = 20  # Установите нужную ширину  столбца
    worksheet.column_dimensions['H'].width = 42  # Установите нужную ширину  столбца
    worksheet.column_dimensions['I'].width = 47  # Установите нужную ширину  столбца

    # Форматирование заголовков
    blue_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')  # цвет
    green_fill = PatternFill(start_color='d9f2dd', end_color='d9f2dd', fill_type='solid')  # цвет
    yellou_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # цвет

    # Индексы заголовков, к которым нужно применить зеленый фон (начиная с 1)
    blue_header_indices = [1, 2, 3, 4, 5, 6, 7, 8, 9]  # Индексы для 'ФИО врача',....', 'Вывод'
    green_header_indices = [10,11,12,13, 14, 15,16,17,18,19,20,
                            21,22,23,24,25,26,27,28,29,30,31]  # Индексы для 'ФИО врача',....', 'Вывод'

    yellou_header_indices = [32]  # Индексы для '...'

    for col in blue_header_indices:
        worksheet.cell(row=1, column=col).fill = blue_fill  # Применяем зеленый фон к указанным заголовкам

    for col in green_header_indices:
        worksheet.cell(row=1, column=col).fill = green_fill  # Применяем зеленый фон к указанным заголовкам

    for col in yellou_header_indices:
        worksheet.cell(row=1, column=col).fill = yellou_fill  # Применяем зеленый фон к указанным заголовкам

    # Установка переноса текста для заголовков и жирного шрифта
    bold_font = Font(bold=True)  # Создаем объект шрифта с жирным начертанием

    # Создание границ
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    for col in range(1, len(headers) + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.alignment = Alignment(wrap_text=True, horizontal='center',
                                   vertical='center')  # Устанавливаем перенос текста
        cell.font = bold_font  # Устанавливаем жирный шрифт
        cell.border = thin_border  # Применяем границы к ячейке заголовка

        # Заполнение данными
    # Заполнение данными
    for item in data_plf:
        row_data = [
            item.data_planfakta.strftime("%d.%m.%Y") if item.data_planfakta else '',  # Форматируем дату
            item.id_pac,
            item.polis_oms,
            item.fio_pacienta,
            item.data_rozhdeniya.strftime("%d.%m.%Y") if item.data_rozhdeniya else '',  # Форматируем дату
            item.ovpp_name,
            item.planovaya_data_vizita_soglasno_emias_reestru.strftime(
                "%d.%m.%Y") if item.planovaya_data_vizita_soglasno_emias_reestru else '',  # Форматируем дату
            item.fakticheskaya_data_vizita_soglasno_emias.strftime(
                "%d.%m.%Y") if item.fakticheskaya_data_vizita_soglasno_emias else '',  # Форматируем дату
            item.vopros_dlya_razbora,

            # Форматируем дату
            item.data_vk.strftime(
                "%d.%m.%Y") if item.data_vk else '',  # Форматируем дату ,#'Дата ВК (ОК)',

            item.name_mo_provod_vk,#'Наименование МО, проводившей ВК(ОК)',
            item.diagnoz_mkb10,#'Диагноз МКБ-10',
            # Форматируем дату
            item.data_vklyucheniya_v_reestr.strftime(
                "%d.%m.%Y") if item.data_vklyucheniya_v_reestr else '',   #Дата внесения в Реестр',
            # Форматируем дату
            item.data_pervichnogo_vizita.strftime(
                "%d.%m.%Y") if item.data_pervichnogo_vizita else '',   #'Дата первичного визита',

            item.komment_k_pervichnomu_vizitu,#'Комментарий к первичному визиту(при необходимости)',
            # Форматируем дату
            item.data_poslednego_vizita.strftime(
                "%d.%m.%Y") if item.data_poslednego_vizita else '',  #  #'Дата последнего визита',

            item.komment_k_posled_vizitu,#'Комментарий к последнему визиту(при необходимости)',
            item.tyagostnaya_simptomatika_pall_potrebnosti,#'Имеющаяся тягостная симптоматика, паллиативные потребности',
            item.plan_nablyudeniya_sootvestvuet_tyazhesti_sostoyaniya_i_prognozu,#'Индивидуальный план наблюдения соотвествует тяжести состояния и прогнозу',
            item.vizov_smp_posle_last_vizit,#'Вызовы СМП после последнего визита(даты)',
            item.nalichie_pokazanij_k_okazaniyu_specPMP,#'Наличие показаний к оказанию специализированной ПМП',
            item.nablyudenie_v_drugoj_mo_parallel_s_ovpp,#'Наблюдение в других МО (параллельно с ОВППМП)',
            item.name_drugoj_mo_parallel_s_ovpp,#'Наименование другой МО, в которой наблюдается пациент параллельно с параллельно с ОВППМП',
            # Форматируем дату
            item.data_poslednego_poseshcheniya_drugoj_mo.strftime(
                "%d.%m.%Y") if item.data_poslednego_poseshcheniya_drugoj_mo else '',  #'Дата последнего посещения другой МО',

            item.osnovaniya_parallel_nabluden,#'Обоснование параллельного наблюдения иными МО',
            item.vypiska_receptov_vps,#'Выписка рецептов для лечения тягостной симптматики ОВППМП дата последней выписки, наименование препарата, на какой срок или количество препарата',
            item.vypiska_receptov_drugoj_mo,#'Выписка рецептов для лечения тягостной симптматикидругой МО дата, наименование препарата, на какой срок или количество препарата',
            item.potrebnost_v_respiratorke,#,'Потребность в респираторной поддержке да / нет, при наличии потребности - комментарии, в том числе дата выявления потребности и дата обеспечения респираторным оборудованием',
            item.vyvlennye_defekty,#'Дефекты, выявленные в процессе разбора',
            item.mery_prinyatye_vps_dlya_uluchsheniya_pmp,#'Меры, принятые в ОВППМП для оказания качественной ПМП',
            item.predlozheniya_dlya_uluchsheniya_pmp,#'Предложения по организационным решениям для улучшения качества оказания ПМП',
            item.otvet_kc, #'Коментарий ЦПП',
        ]

        # Добавляем данные в строку
        worksheet.append(row_data)

        # Применяем границы ко всем ячейкам в текущей строке
        current_row = worksheet.max_row  # Получаем номер последней строки
        for col in range(1, len(row_data) + 1):
            cell = worksheet.cell(row=current_row, column=col)
            cell.border = thin_border  # Применяем границы к ячейке


    # Сохранение файла в памяти
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    # ----- Формирую название файла --------
    if search_data_zamecaniya:
        file = f'plan_fact_{search_data_zamecaniya}.xlsx'
    else:      file = 'plan_fact_all.xlsx'
    print(file)

    # Создание HTTP ответа
    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{file}"'
    return response