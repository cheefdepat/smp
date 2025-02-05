from django.shortcuts import render, get_object_or_404, redirect
from .models import SmpRazborTab
from django.db import connection
from django.core.paginator import Paginator
from django import forms
import datetime
from django.contrib.auth import authenticate, login as auth_login
from django.contrib.auth.views import LoginView, LogoutView
from django.urls import reverse, reverse_lazy
from django.contrib.auth import authenticate, login as auth_login, logout as auth_logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import Group
import logging

import openpyxl
from io import BytesIO
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from django.http import HttpResponse

# ------------------------------------------------- Зона логгирования ----------------
class MyFilter(logging.Filter):
    def filter(self, record):
        return not record.getMessage().startswith('GET') and not record.getMessage().startswith('POST')

# Создайте логгер
logger = logging.getLogger(__name__)

# Установите уровень логирования
logger.setLevel(logging.INFO)

# Создайте обработчик логирования для файла
file_handler = logging.FileHandler('log_smp.txt')

# Установите формат логирования
# formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
# formatter = logging.Formatter('%(asctime)s - %(name)s -  %(message)s')
# file_handler.setFormatter(formatter)
# Добавьте фильтр логирования к обработчику логирования
file_handler.addFilter(MyFilter())

# Добавьте обработчик логирования к логгеру
logger.addHandler(file_handler)

# -------------------------------------------------  логгирования -----------------------

def login_view(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)

        if user is not None:
            auth_login(request, user)
            print(f"-----------------------------------------------------User {user} вошел")
            logger.info(f'пользователь {request.user} вошел в приложение')
            return redirect('my_app_smp:start_page')
            # return redirect('start_page')
        else:
            error_message = "Вас нет в списке"
            return render(request, 'login.html', {'error_message': error_message})

    return render(request, 'login.html')


def logout(request):
    auth_logout(request)  # Выход из аккаунта
    logger.info(f'пользователь {request.user} вышел из приложения')
    print("-----------------------------------------00-----User logged out")  # Отладочное сообщение
    return redirect('login')  # Перенаправление на главную страницу
    # return render(request, 'login.html')

def start_page(request):
    # -------------- получаю принадлежность к группе -------
    user_groups_list = []
    for i in request.user.groups.all():
        print(i)
        user_groups_list.append(str(i))

    return render(request, 'start.html',
                  {
                      'groups': user_groups_list,
                         }              )

def ny(request):
    return render(request, 'ny.html')

# class MyprojectLogout(LogoutView):
#     next_page  = reverse_lazy('login')

def zamena_pustot(pole_proverki_daty):
    # Функция проверки является ли значение в поле ДАТЫ - пустым??? для ВСЕХ ДАТ проверку!!!!
    if pole_proverki_daty == "":
        pole_proverki_daty = None
    else:
        pole_proverki_daty = pole_proverki_daty
    return pole_proverki_daty

def calculate_date(chislo1, chislo2):
    #--------------- функция вычисления расчетных значений ДАТ:
    if chislo1 and chislo2:
        result = chislo1 - chislo2
    else:
        result = 'нет даты'
    return result

def help(request):
    return render(request, 'help.html', {

    })


def results(request):
    return render(request, 'results.html')

def results_page_1(request):
    return render(request, 'results_page_1.html')

def results_page_2(request):
    return render(request, 'results_page_2.html')

def home(request):

    # --------------
    user_groups_list = []
    for i in request.user.groups.all():
        print(i)
        user_groups_list.append(str(i))
    # --------------

    # ----------------------- отбор списка ВПС
    unique_kurir = SmpRazborTab.objects.values_list('kuriruyushchee_podrazdelenie_ovpp', flat=True).distinct()
    unique_kurir_list = []
    for i in unique_kurir:
        print(i)
        unique_kurir_list.append(str(i))
    print()
    # -----------------------

    is_vps_group = request.user.groups.filter(name='vps').exists()
    query_fio = request.GET.get('search_fio', '')  # Получаем строку поиска по FIO
    query_kurir = request.GET.get('search_kurir', '')  # Получаем строку поиска по курирующему подразделению
    query_otrabot = request.GET.get('search_otrabot', '')  # Получаем строку поиска по отработанным записям
    query_data_vizova_smp = request.GET.get('search_data_vizova_smp', '')  # Получаем строку поиска по дате смп

    records_per_page = request.GET.get('records_per_page', 20)  # Получаем количество записей на странице

    # Обработка фильтров из GET-запроса

    if request.method == 'GET':
        if 'records_per_page' in request.GET and request.GET['records_per_page'] != '':
            request.session['records_per_page'] = int(request.GET['records_per_page'])
        records_per_page = request.GET.get('records_per_page', request.session.get('records_per_page', 15))

        if 'search_fio' in request.GET:
            query_fio = request.GET.get('search_fio', '')

        if 'search_kurir' in request.GET:
            query_kurir = request.GET.get('search_kurir', '')

        if 'search_otrabot' in request.GET:
            query_otrabot = request.GET.get('search_otrabot', '')

        if 'search_data_vizova_smp' in request.GET:
            query_data_vizova_smp = request.GET.get('search_data_vizova_smp', '')
        # -----------------------------------------------------------


    # Фильтруем данные по обоим полям и сортируем по p_p
    # data_smp = SmpRazborTab.objects.all().order_by('p_p')  # Сортировка по возрастанию p_p
    # Фильтруем данные по обоим полям
    if user_groups_list == ['vps']:
        # Исключаем записи, у которых ok_vps равно "ВПС"
        data_smp = SmpRazborTab.objects.filter(ok_vps="впс").order_by('data_vyzova_smp',
                                                                                  'fio_pacienta')  # Сортировка по возрастанию
    else:
        data_smp = SmpRazborTab.objects.all().order_by('data_vyzova_smp', 'fio_pacienta')  # Сортировка по возрастанию

    total_records = data_smp.count()  # Общее количество записей

    if query_fio:
        data_smp = data_smp.filter(fio_pacienta__icontains=query_fio)

    if query_kurir:
        data_smp = data_smp.filter(kuriruyushchee_podrazdelenie_ovpp__icontains=query_kurir)

    if query_otrabot:
        data_smp = data_smp.filter(ok_vps=query_otrabot)

    if query_data_vizova_smp:
       data_smp = data_smp.filter(data_vyzova_smp=query_data_vizova_smp)


    # Получаем уникальные значения для выпадающего списка
    unique_kurir = SmpRazborTab.objects.values_list('kuriruyushchee_podrazdelenie_ovpp', flat=True).distinct()
    unique_otrab = SmpRazborTab.objects.values_list('ok_vps', flat=True).distinct()

    paginator = Paginator(data_smp, records_per_page)  # Показывать 10 записей на странице
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    filter_records = data_smp.count()  # количество записей - отфильтрованное

    return render(request, 'home.html', {
        'data_smp': page_obj,
        'paginator': paginator,

        'search_fio': query_fio,
        'search_kurir': query_kurir,
        'search_otrabot': query_otrabot,
        'search_data_vizova_smp': query_data_vizova_smp,

        'records_per_page': records_per_page,
        'total_records': total_records,  # Передаем общее количество записей
        'unique_kurir': unique_kurir,  # Передаем уникальные значения в контекст по курир. филиалу ВПС
        'unique_otrab': unique_otrab,  # Передаем уникальные значения в контекст по отработанным в КЭР
        'groups': user_groups_list, # Получаем все группы
        'filter_records': filter_records, ## количество записей - отфильтрованное
    })



# @login_required
def patient_detail(request, id):
    patient = get_object_or_404(SmpRazborTab, id=id)  # Получаем запись по ID

    return render(request, 'patient_detail.html', {'patient': patient})


def edit_patient(request, id):
    patient = get_object_or_404(SmpRazborTab, id=id)

    # --------------- вычсление перес=менных по ДНЯМ!!! ------------

    koli4_dney_ot_proshlogo_vizita_vracha = calculate_date(patient.kakaya_data_sleduyushchego_vizita_vracha_soglasno_protokolu,
                                                           patient.data_poslednego_vizita_vracha_iz_protokola_osmotra_emias)
    dni_kolichestvo_dnej_ot_momenta_polucheniya_dannykh_po_smp_do_sover = calculate_date(patient.data_polucheniya_svedenij_po_vyzovam_smp_ot_kc,
                                                           patient.data_naznachenogo_audioprotokola_soglasno_protokolu_v)

    if request.method == 'POST':
        # Обновляем поля, которые можно редактировать
        # ------------------- поля для редактирования и сохранения --------

        patient.fio_vracha = request.POST.get('fio_vracha')
        patient.sravnenie_povoda_vyz_smp_s_prot_vracha_cpp_do_i_posle = request.POST.get(
            'sravnenie_povoda_vyz_smp_s_prot_vracha_cpp_do_i_posle')
        patient.dejstviya_cpp = request.POST.get('dejstviya_cpp')
        patient.kontrol_ispolneniya_naznachennykh_vrachom_cpp_rekomendacij = request.POST.get(
            'kontrol_ispolneniya_naznachennykh_vrachom_cpp_rekomendacij')

        # -----------------------------Данные из последнего протокола врача --------------------------
        patient.byl_li_ustanovlen_bazovyj_plan_na_poslednem_vizite = request.POST.get(
            'byl_li_ustanovlen_bazovyj_plan_na_poslednem_vizite')
        patient.kakaya_data_sleduyushchego_vizita_vracha_soglasno_protokolu = request.POST.get(
            'kakaya_data_sleduyushchego_vizita_vracha_soglasno_protokolu')

        # Проверяем, является ли значение None ---------------- для ВСЕХ ДАТ проверку!!!!
        patient.kakaya_data_sleduyushchego_vizita_vracha_soglasno_protokolu = zamena_pustot(patient.kakaya_data_sleduyushchego_vizita_vracha_soglasno_protokolu)



        # -----------------
        patient.otobrazheny_li_vse_zhaloby_pacienta_v_polnom_obeme = request.POST.get(
            'otobrazheny_li_vse_zhaloby_pacienta_v_polnom_obeme')
        patient.dinamika_sostoyaniya = request.POST.get('dinamika_sostoyaniya')
        patient.data_naznachenogo_audioprotokola_soglasno_protokolu_v = request.POST.get(
            'data_naznachenogo_audioprotokola_soglasno_protokolu_v')

        # Проверяем ДАТУ, является ли значение None ---------------- для ВСЕХ ДАТ проверку!!!!
        patient.data_naznachenogo_audioprotokola_soglasno_protokolu_v = zamena_pustot(
            patient.data_naznachenogo_audioprotokola_soglasno_protokolu_v)

        # -----------------------------Проверка заведующего --------------------------

        # --------------------Проверяем ДАТУ !!!!--------------------------------------------------
        # -----Начало блока проверки даты------ получение ДАТЫ и ее проверка на пустоту!!!--------
        patient.data_polucheniya_svedenij_po_vyzovam_smp_ot_kc = request.POST.get(
            'data_polucheniya_svedenij_po_vyzovam_smp_ot_kc')
        patient.data_polucheniya_svedenij_po_vyzovam_smp_ot_kc = zamena_pustot(
            patient.data_polucheniya_svedenij_po_vyzovam_smp_ot_kc)
        # ---------Конец блока проверки даты- получение ДАТЫ и ее проверка на пустоту!!!-----------

        # -----Начало блока проверки даты------ получение ДАТЫ и ее проверка на пустоту!!!--------
        patient.data_audioprotokola_posle_polucheniya_dannykh_o_vyzove_smp = request.POST.get(
            'data_audioprotokola_posle_polucheniya_dannykh_o_vyzove_smp')
        patient.data_audioprotokola_posle_polucheniya_dannykh_o_vyzove_smp = zamena_pustot(
            patient.data_audioprotokola_posle_polucheniya_dannykh_o_vyzove_smp)
        # ---------Конец блока проверки даты- получение ДАТЫ и ее проверка на пустоту!!!-----------


        patient.kakova_prichina_vyzova_smp_po_rezutatam_audiokontrolya = request.POST.get(
            'kakova_prichina_vyzova_smp_po_rezutatam_audiokontrolya')
        patient.kakie_dejstviya_byli_predprinyaty_smp = request.POST.get('kakie_dejstviya_byli_predprinyaty_smp')
        patient.tekushchee_sostoyanie_pacienta_posle_vyzova_smp = request.POST.get(
            'tekushchee_sostoyanie_pacienta_posle_vyzova_smp')
        patient.ostalis_li_zhaloby_posle_vyzova_smp = request.POST.get('ostalis_li_zhaloby_posle_vyzova_smp')
        patient.opisanie_zhalob = request.POST.get('opisanie_zhalob')
        patient.byli_li_lekarstvennye_sredstva_otovareny_po_receptu = request.POST.get(
            'byli_li_lekarstvennye_sredstva_otovareny_po_receptu')
        patient.pacient_prinimaet_naznachennye_lekarstvennye_sredstva = request.POST.get(
            'pacient_prinimaet_naznachennye_lekarstvennye_sredstva')
        patient.ehffektivna_li_naznachennaya_medikamentoznaya_terapiya = request.POST.get(
            'ehffektivna_li_naznachennaya_medikamentoznaya_terapiya')
        patient.est_li_pobochnye_dejstviya_ot_naznachennykh_lekarstvennykh = request.POST.get(
            'est_li_pobochnye_dejstviya_ot_naznachennykh_lekarstvennykh')
        patient.pacientu_byla_ozvuchena_data_sleduyushchego_vizita = request.POST.get(
            'pacientu_byla_ozvuchena_data_sleduyushchego_vizita')
        patient.pacientu_byli_predostavleny_kontaktnye_nomera_cpp = request.POST.get(
            'pacientu_byli_predostavleny_kontaktnye_nomera_cpp')
        patient.pacient_zvonil_po_ukazannym_kontaktnym_nomera_do_vyzova_smp = request.POST.get(
            'pacient_zvonil_po_ukazannym_kontaktnym_nomera_do_vyzova_smp')
        patient.prichina_po_kotoroj_pacient_ne_zvonil_po_ukazannym_nomeram = request.POST.get(
            'prichina_po_kotoroj_pacient_ne_zvonil_po_ukazannym_nomeram')
        patient.sootvetstvuet_li_naznachennyj_bazovyj_plan_sostoyaniyu_pacie = request.POST.get(
            'sootvetstvuet_li_naznachennyj_bazovyj_plan_sostoyaniyu_pacie')
        patient.trebovanie_gospitalizacii_na_dannyj_moment = request.POST.get(
            'trebovanie_gospitalizacii_na_dannyj_moment')
        patient.vyyavlennye_defekty_v_rabote_vracha = request.POST.get('vyyavlennye_defekty_v_rabote_vracha')

        # -----------------------

        print('1ssssss-----------------')
        if 'save' in request.POST:  # Кнопка "Сохранить"
            print('ssssss-----------------')
            patient.kolichestvo_dnej_ot_proshlogo_vizita_vracha = koli4_dney_ot_proshlogo_vizita_vracha
            patient.kolichestvo_dnej_ot_momenta_polucheniya_dannykh_po_smp_do_sover = dni_kolichestvo_dnej_ot_momenta_polucheniya_dannykh_po_smp_do_sover
            logger.info(f'пользователь {request.user} сохранил запись о пациенте {patient.fio_pacienta}')
            patient.save()
            return redirect('my_app_smp:home')

        elif 'send_to_ker' in request.POST:  # Кнопка "Отправить в КЭР"
            # if patient.is_valid():
            logger.info(f'пользователь {request.user} отправил запись о пациенте {patient.fio_pacienta} в КЭР')
            patient.kolichestvo_dnej_ot_proshlogo_vizita_vracha = koli4_dney_ot_proshlogo_vizita_vracha
            patient.kolichestvo_dnej_ot_momenta_polucheniya_dannykh_po_smp_do_sover = dni_kolichestvo_dnej_ot_momenta_polucheniya_dannykh_po_smp_do_sover

            patient.save()  # Сохраняем изменения
            return redirect('my_app_smp:proverka', id=patient.id)  # Переходим на страницу проверки

        # patient.save()  # Сохраняем изменения
        # return redirect('home',
        #
        #                 )  # Перенаправляем на главную страницу

    return render(request, 'edit_pacient_short.html', {'patient': patient,
                                                       'koli4_dney_ot_proshlogo_vizita_vracha': koli4_dney_ot_proshlogo_vizita_vracha,
                                                       'dni_kolichestvo_dnej_ot_momenta_polucheniya_dannykh_po_smp_do_sover': dni_kolichestvo_dnej_ot_momenta_polucheniya_dannykh_po_smp_do_sover,
                                                       })

def edit_ker(request, id):
    patient = get_object_or_404(SmpRazborTab, id=id)
    #  ----- отображение рассчетных ДННЕЙ+++++++--------------------
    koli4_dney_ot_proshlogo_vizita_vracha = calculate_date(patient.kakaya_data_sleduyushchego_vizita_vracha_soglasno_protokolu,
                                                           patient.data_poslednego_vizita_vracha_iz_protokola_osmotra_emias)
    dni_kolichestvo_dnej_ot_momenta_polucheniya_dannykh_po_smp_do_sover = calculate_date(patient.data_polucheniya_svedenij_po_vyzovam_smp_ot_kc,
                                                           patient.data_naznachenogo_audioprotokola_soglasno_protokolu_v)

    if request.method == 'POST':
        # --------------- поля для редактирования и сохранения --------
        patient.bazovyj_plan_naznachen_korrektno = request.POST.get('bazovyj_plan_naznachen_korrektno')
        patient.ocenka_sostoyaniya_sootvetstvuet_bazovomu_planu = request.POST.get('ocenka_sostoyaniya_sootvetstvuet_bazovomu_planu')
        patient.zhaloby_opisany_v_polnom_obeme = request.POST.get('zhaloby_opisany_v_polnom_obeme')
        patient.ocenka_zaveduyushchego_proizvedena_korrektno = request.POST.get('ocenka_zaveduyushchego_proizvedena_korrektno')
        patient.ocenka_dejstvij_vracha_do_vyzova_smp = request.POST.get('ocenka_dejstvij_vracha_do_vyzova_smp')
        patient.ocenka_dejstvij_posle_vyzova_smp = request.POST.get('ocenka_dejstvij_posle_vyzova_smp')
        patient.vyvody_po_rezultatm_ocenki = request.POST.get('vyvody_po_rezultatm_ocenki')

        # patient.save()  # Сохраняем изменения

        print('1ker-----------------')
        if 'save_ker' in request.POST:  # Кнопка "Сохранить"
            patient.kolichestvo_dnej_ot_proshlogo_vizita_vracha = koli4_dney_ot_proshlogo_vizita_vracha
            patient.kolichestvo_dnej_ot_momenta_polucheniya_dannykh_po_smp_do_sover = dni_kolichestvo_dnej_ot_momenta_polucheniya_dannykh_po_smp_do_sover

            patient.save()
            return redirect('my_app_smp:home')

        elif 'return_na_vps' in request.POST:  # Кнопка "Отправить в КЭР"
            patient.ok_vps = "впс"
            patient.kolichestvo_dnej_ot_proshlogo_vizita_vracha = koli4_dney_ot_proshlogo_vizita_vracha
            patient.kolichestvo_dnej_ot_momenta_polucheniya_dannykh_po_smp_do_sover = dni_kolichestvo_dnej_ot_momenta_polucheniya_dannykh_po_smp_do_sover

            patient.save()  # Сохраняем изменения
            return redirect('my_app_smp:home')  # Переходим на страницу проверки

        elif 'go_gv' in request.POST:  # Кнопка "Отправить в КЭР"
            patient.ok_vps = "Передано Глав.врачу"
            patient.save()  # Сохраняем изменения
            return redirect('my_app_smp:proverka_ker', id=patient.id)  # Переходим на страницу проверки

        return redirect('my_app_smp:home',

                        )  # Перенаправляем на главную страницу

    return render(request, 'edit_pacient_ker.html', {'patient': patient,
                                                       'koli4_dney_ot_proshlogo_vizita_vracha': koli4_dney_ot_proshlogo_vizita_vracha,
                                                       'dni_kolichestvo_dnej_ot_momenta_polucheniya_dannykh_po_smp_do_sover':dni_kolichestvo_dnej_ot_momenta_polucheniya_dannykh_po_smp_do_sover,
                                                       })


def proverka(request, id):
    patient = get_object_or_404(SmpRazborTab, id=id)
    # --------------- вычсление перес=менных по ДНЯМ!!! ------------
    koli4_dney_ot_proshlogo_vizita_vracha = calculate_date(patient.kakaya_data_sleduyushchego_vizita_vracha_soglasno_protokolu,
                                                           patient.data_poslednego_vizita_vracha_iz_protokola_osmotra_emias)
    dni_kolichestvo_dnej_ot_momenta_polucheniya_dannykh_po_smp_do_sover = calculate_date(patient.data_polucheniya_svedenij_po_vyzovam_smp_ot_kc,
                                                           patient.data_naznachenogo_audioprotokola_soglasno_protokolu_v)

    if request.method == 'POST':
        if 'ot_vps_v_ker' in request.POST:  # Кнопка "Отправить в КЭР"
            patient.ok_vps = "Передано в КЭР"
            patient.save()
            return redirect('my_app_smp:home')  # Переходим на главную страницу

        elif 'korrektirovat' in request.POST:  # Кнопка "Корректировать"
            return redirect('my_app_smp:edit_patient', id=id)  # Возвращаем на страницу редактирования

    return render(request, 'patient_detail.html', {'patient': patient,
                                                       'koli4_dney_ot_proshlogo_vizita_vracha': koli4_dney_ot_proshlogo_vizita_vracha,
                                                       'dni_kolichestvo_dnej_ot_momenta_polucheniya_dannykh_po_smp_do_sover':dni_kolichestvo_dnej_ot_momenta_polucheniya_dannykh_po_smp_do_sover,
                                                       })



def proverka_ker(request, id):
    patient = get_object_or_404(SmpRazborTab, id=id)
    if request.method == 'POST':
        if 'ot_ker_na_glav' in request.POST:  # Кнопка "Отправить в КЭР"
            patient.ok_vps = "Передано Глав.врачу"
            patient.save()
            return redirect('my_app_smp:home')  # Переходим на главную страницу

        elif 'korrektirovat_ker' in request.POST:  # Кнопка "Корректировать"
            patient.ok_vps = "Передано в КЭР"
            patient.save()
            return redirect('my_app_smp:home')  # Возвращаем на страницу редактирования

    return render(request, 'proverka_ot_ker_na_glav.html', {'patient': patient})

def edit_glav(request, id):
    patient = get_object_or_404(SmpRazborTab, id=id)
    #  ----- отображение рассчетных ДННЕЙ+++++++--------------------
    koli4_dney_ot_proshlogo_vizita_vracha = calculate_date(patient.kakaya_data_sleduyushchego_vizita_vracha_soglasno_protokolu,
                                                           patient.data_poslednego_vizita_vracha_iz_protokola_osmotra_emias)
    dni_kolichestvo_dnej_ot_momenta_polucheniya_dannykh_po_smp_do_sover = calculate_date(patient.data_polucheniya_svedenij_po_vyzovam_smp_ot_kc,
                                                           patient.data_naznachenogo_audioprotokola_soglasno_protokolu_v)




    if request.method == 'POST':
        # Обновляем поля, которые можно редактировать и СОХРАНЯТЬ!!!
        # 'analiz_dejstvij_glavnym_vrachom',

        # --------------- поля для редактирования и сохранения --------
        patient.analiz_dejstvij_glavnym_vrachom = request.POST.get('analiz_dejstvij_glavnym_vrachom')
        patient.vyvod = request.POST.get('vyvod')


        # patient.save()  # Сохраняем изменения

        print('1ker-----------------')
        if 'save_glav' in request.POST:  # Кнопка "Сохранить"
            # if patient.is_valid():
            patient.save()
            return redirect('my_app_smp:home')

        elif 'return_na_vps' in request.POST:  # Кнопка "Отправить в КЭР"
            patient.ok_vps = "впс"
            patient.save()  # Сохраняем изменения
            return redirect('my_app_smp:home')  # Переходим на страницу проверки

        elif 'return_na_ker' in request.POST:  # Кнопка "Отправить в КЭР"
            patient.ok_vps = "Передано КЭР"
            patient.save()  # Сохраняем изменения
            return redirect('my_app_smp:home')  # Переходим на страницу проверки

        elif 'go_dzm' in request.POST:  # Кнопка "Отправить в КЭР"
            patient.ok_vps = "Передано ДЗМ"
            patient.save()  # Сохраняем изменения
            return redirect('my_app_smp:proverka_glav', id=patient.id)  # Переходим на страницу проверки

        return redirect('my_app_smp:home',

                        )  # Перенаправляем на главную страницу

    return render(request, 'edit_pacient_glav.html', {'patient': patient,
                                                       'koli4_dney_ot_proshlogo_vizita_vracha': koli4_dney_ot_proshlogo_vizita_vracha,
                                                       'dni_kolichestvo_dnej_ot_momenta_polucheniya_dannykh_po_smp_do_sover':dni_kolichestvo_dnej_ot_momenta_polucheniya_dannykh_po_smp_do_sover,
                                                       })

def proverka_glav(request, id):
    patient = get_object_or_404(SmpRazborTab, id=id)
    if request.method == 'POST':
        if 'ot_glav_na_dzm' in request.POST:  # Кнопка "Отправить в КЭР"
            patient.ok_vps = "Передано ДЗМ"
            patient.save()
            return redirect('my_app_smp:home')  # Переходим на главную страницу

        elif 'korrektir_glav' in request.POST:  # Кнопка "Корректировать"
            patient.ok_vps = "Передано Глав.врачу"
            patient.save()
            return redirect('my_app_smp:home')  # Возвращаем на страницу редактирования

    return render(request, 'proverka_ot_glav.html', {'patient': patient})

def custom_404_view(request, exception):
    return render(request, '404.html', status=404)

def custom_500_view(request):
    return render(request, '500.html', status=500)

def export_to_excel(request):
    # Получите данные с учетом фильтров
    # search_fio = request.GET.get('search_fio', '')
    # search_kurir = request.GET.get('search_kurir', '')
    # search_otrabot = request.GET.get('search_otrabot', '')
    search_data_vizova_smp = request.GET.get('search_data_vizova_smp', '')

    # Примените фильтры к вашему запросу
    data_smp = SmpRazborTab.objects.all()  # Замените на ваш запрос
    # if search_fio:
    #     data_smp = data_smp.filter(fio_pacienta__icontains=search_fio)
    # if search_kurir:
    #     data_smp = data_smp.filter(kuriruyushchee_podrazdelenie_ovpp=search_kurir)
    # if search_otrabot:
    #     data_smp = data_smp.filter(status_otrabotki=search_otrabot)  # Замените на ваше поле
    if search_data_vizova_smp:
        data_smp = data_smp.filter(data_vyzova_smp=search_data_vizova_smp)  # Замените на ваше поле

    # Создание Excel файла
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Patients Data'

    # Заголовки столбцов
    headers = [
        'Дата вызова СМП',
        'ФИО пациента',
        'Количество выездов в день',
        'Дата рождения',
        'Полис ОМС',
        'Результат вызова',
        'Актив / Пассив',
        'Диагноз по МКБ',
        'Причина вызова скорой помощи ', # -------------
        'Наличие болевого синдрома',
        'Причина вызова скорой помощи (кратко)',
        'Курирующее подразделение ОВПП',
        'Дата включения в регистр',
        'Дата последнего визита врача из протокола осмотра EMIAS',
        'ФИО врача',
        'Сравнение "Повода вызова СМП" с Протоколами врача ЦПП выездов ДО и ПОСЛЕ вызова СМП',
        'Действия ЦПП',
        'Контроль исполнения назначенных врачом ЦПП рекомендаций',
        'Вывод',
    ]
    worksheet.append(headers)

    # Установка ширины для 2-го и 15-го столбца
    worksheet.column_dimensions['A'].width = 10  # Установите нужную ширину для 2-го столбца
    worksheet.column_dimensions['B'].width = 38  # Установите нужную ширину для 2-го столбца
    worksheet.column_dimensions['C'].width = 15  # Установите нужную ширину столбца
    worksheet.column_dimensions['D'].width = 15  # Установите нужную ширину  столбца
    worksheet.column_dimensions['E'].width = 20  # Установите нужную ширину  столбца
    worksheet.column_dimensions['F'].width = 25  # Установите нужную ширину  столбца
    worksheet.column_dimensions['G'].width = 15  # Установите нужную ширину  столбца
    worksheet.column_dimensions['H'].width = 15  # Установите нужную ширину  столбца
    worksheet.column_dimensions['I'].width = 35  # Установите нужную ширину  столбца
    worksheet.column_dimensions['J'].width = 15  # Установите нужную ширину  столбца
    worksheet.column_dimensions['K'].width = 20  # Установите нужную ширину  столбца
    worksheet.column_dimensions['L'].width = 24  # Установите нужную ширину  столбца
    worksheet.column_dimensions['M'].width = 20  # Установите нужную ширину  столбца
    worksheet.column_dimensions['N'].width = 20  # Установите нужную ширину  столбца
    worksheet.column_dimensions['O'].width = 40  # Установите нужную ширину  столбца
    worksheet.column_dimensions['P'].width = 40  # Установите нужную ширину  столбца
    worksheet.column_dimensions['Q'].width = 40  # Установите нужную ширину  столбца
    worksheet.column_dimensions['R'].width = 40  # Установите нужную ширину  столбца
    worksheet.column_dimensions['S'].width = 40  # Установите нужную ширину  столбца

    # Форматирование заголовков
    grey_fill = PatternFill(start_color='BFBFBF', end_color='BFBFBF', fill_type='solid')  # Зеленый цвет
    yellou_fill = PatternFill(start_color='FDE9D9', end_color='FDE9D9', fill_type='solid')  # Зеленый цвет
    green_puff_fill = PatternFill(start_color='D8E4BC', end_color='D8E4BC', fill_type='solid')  # Зеленый цвет

    # Индексы заголовков, к которым нужно применить зеленый фон (начиная с 1)
    grey_header_indices = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10]  # Индексы для 'ФИО врача',....', 'Вывод'
    yellou_header_indices = [11, 12, 13, 14]  # Индексы для '...'
    green_puff_indices = [15, 16, 17, 18, 19]  # Индексы для 'ФИО врача',....', 'Вывод'

    for col in grey_header_indices:
        worksheet.cell(row=1, column=col).fill = grey_fill  # Применяем зеленый фон к указанным заголовкам

    for col in yellou_header_indices:
        worksheet.cell(row=1, column=col).fill = yellou_fill  # Применяем зеленый фон к указанным заголовкам

    for col in green_puff_indices:
        worksheet.cell(row=1, column=col).fill = green_puff_fill  # Применяем зеленый фон к указанным заголовкам

    # Установка переноса текста для заголовков и жирного шрифта
    bold_font = Font(bold=True)  # Создаем объект шрифта с жирным начертанием

    # Создание границ
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    for col in range(1, len(headers) + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.alignment = Alignment(wrap_text=True, horizontal='center',
                                   vertical='center')  # Устанавливаем перенос текста
        cell.font = bold_font  # Устанавливаем жирный шрифт
        cell.border = thin_border  # Применяем границы к ячейке заголовка

        # Заполнение данными

    for item in data_smp:
        row_data = [
            item.data_vyzova_smp.strftime("%d.%m.%Y") if item.data_vyzova_smp else '',  #
            item.fio_pacienta,  #
            item.kolichestvo_vyezdov_v_den,  #
            item.data_rozhdeniya.strftime("%d.%m.%Y") if item.data_rozhdeniya else '',  #
            item.polis_oms,
            item.rezultat_vyzova,
            item.aktiv_passiv,
            item.diagnoz_po_mkb,
            item.prichina_vyzova_skoroiy_pomoschi,
            item.nalichie_bolevogo_sindroma,
            item.prichina_vyzova_skoroiy_pomoschi_kratko,
            item.kuriruyushchee_podrazdelenie_ovpp,
            item.data_vklyucheniya_v_registr.strftime("%d.%m.%Y") if item.data_vklyucheniya_v_registr else '',
            item.data_poslednego_vizita_vracha_iz_protokola_osmotra_emias.strftime(
                "%d.%m.%Y") if item.data_poslednego_vizita_vracha_iz_protokola_osmotra_emias else '',
            item.fio_vracha,
            item.sravnenie_povoda_vyz_smp_s_prot_vracha_cpp_do_i_posle,
            item.dejstviya_cpp,
            item.kontrol_ispolneniya_naznachennykh_vrachom_cpp_rekomendacij,
            item.vyvod,

            # item.kakova_prichina_vyzova_smp_po_rezutatam_audiokontrolya,
            # item.kakie_dejstviya_byli_predprinyaty_smp,
            # item.sootvetstvuet_li_naznachennyj_bazovyj_plan_sostoyaniyu_pacie,
            # item.trebovanie_gospitalizacii_na_dannyj_moment,
            # item.vyyavlennye_defekty_v_rabote_vracha,
            # item.bazovyj_plan_naznachen_korrektno,
            # item.ocenka_sostoyaniya_sootvetstvuet_bazovomu_planu,
            # item.zhaloby_opisany_v_polnom_obeme,
            # item.ocenka_zaveduyushchego_proizvedena_korrektno,
            # item.ocenka_dejstvij_vracha_do_vyzova_smp,
            # item.ocenka_dejstvij_posle_vyzova_smp,
            # item.vyvody_po_rezultatm_ocenki,

        ]

        # Добавляем данные в строку
        worksheet.append(row_data)

        # Применяем границы ко всем ячейкам в текущей строке
        current_row = worksheet.max_row  # Получаем номер последней строки
        for col in range(1, len(row_data) + 1):
            cell = worksheet.cell(row=current_row, column=col)
            cell.border = thin_border  # Применяем границы к ячейке

    # Сохранение файла в памяти
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    # ----- Формирую название файла --------
    if search_data_vizova_smp:
        file = f'smp_{search_data_vizova_smp}.xlsx'
    else:      file = 'smp_all.xlsx'
    print(file)

    # Создание HTTP ответа
    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{file}"'
    return response

def export_to_excel_for_ker(request):
    '''расширенный вывод в ЭКСЕЛЬ данных о СПМ - для КЭР '''

    # Получите данные с учетом фильтров
    search_data_vizova_smp = request.GET.get('search_data_vizova_smp', '')

    # Примените фильтры к вашему запросу
    data_smp = SmpRazborTab.objects.all()  #  запрос
    if search_data_vizova_smp:
        data_smp = data_smp.filter(data_vyzova_smp=search_data_vizova_smp)  # Замените на ваше поле

    # Создание Excel файла
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Patients Data'

    # Заголовки столбцов

    headers = [
            'Дата вызова СМП',
            'ФИО пациента',
            'Количество выездов в день',
            'Дата рождения',
            'Полис ОМС',
            'Результат вызова',
            'Актив / Пассив',
            'Диагноз по МКБ',
            'Причина вызова скорой помощи ', # -------------
            'Наличие болевого синдрома',
            'Причина вызова скорой помощи (кратко)',
            'Курирующее подразделение ОВПП',
            'Дата включения в регистр',
            'Дата последнего визита врача из протокола осмотра EMIAS',
            'ФИО врача',
            'Сравнение "Повода вызова СМП" с Протоколами врача ЦПП выездов ДО и ПОСЛЕ вызова СМП',
            'Действия ЦПП',
            'Контроль исполнения назначенных врачом ЦПП рекомендаций',
            'Вывод',

                'Какова причина вызова СМП по результатам аудиоконтроля?',
                'Какие действия были предприняты СМП?',
                'Соответствует ли назначенный базовый план состоянию пациента',
                'Требуется ли госпитализация на данный момент',
                'Выявленные дефекты в работе врача',
                "Базовый план назначен корректно",
                'Оценка состояния соответствует базовому плану',
                'Жалобы описаны в полном объёме',
                'Оценка заведующего произведена корректно',
                'Оценка действий врача до вызова СМП',
                'Оценка действий после вызова СМП',
                'Выводы по результатм оценки',

        ]


    worksheet.append(headers)


    # Установка ширины для 2-го и 15-го столбца
    worksheet.column_dimensions['B'].width = 30  # Установите нужную ширину для 2-го столбца
    worksheet.column_dimensions['O'].width = 15  # Установите нужную ширину столбца
    worksheet.column_dimensions['P'].width = 20  # Установите нужную ширину  столбца
    worksheet.column_dimensions['Q'].width = 20  # Установите нужную ширину  столбца
    worksheet.column_dimensions['R'].width = 20  # Установите нужную ширину  столбца
    worksheet.column_dimensions['S'].width = 30  # Установите нужную ширину  столбца
    worksheet.column_dimensions['T'].width = 30  # Установите нужную ширину  столбца
    worksheet.column_dimensions['U'].width = 30  # Установите нужную ширину  столбца
    worksheet.column_dimensions['X'].width = 30  # Установите нужную ширину  столбца
    # worksheet.column_dimensions['W'].width = 30  # Установите нужную ширину  столбца
    worksheet.column_dimensions['AE'].width = 30  # Установите нужную ширину  столбца

    # Форматирование заголовков
    green_fill = PatternFill(start_color='98FB98', end_color='98FB98', fill_type='solid')  # Зеленый цвет
    blue_fill = PatternFill(start_color='87CEEB', end_color='87CEEB', fill_type='solid')  # Зеленый цвет
    peach_puff_fill = PatternFill(start_color='FFDAB9', end_color='FFDAB9', fill_type='solid')  # Зеленый цвет
    moccasin_fill = PatternFill(start_color='FFE4B5', end_color='FFE4B5', fill_type='solid')  # Зеленый цвет

    # Индексы заголовков, к которым нужно применить зеленый фон (начиная с 1)
    green_header_indices = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, ]  # Индексы для 'ФИО врача',....', 'Вывод'
    blue_header_indices = [15, 16, 17, 18, 19, ]  # Индексы для '...'

    peach_puff_indices = [20, 21, 22, 23, 24, ]  # Индексы для 'ФИО врача',....', 'Вывод'
    moccasin_fill_indices = [25, 26, 27, 28, 29, 30, 31]  # Индексы для '....'

    for col in green_header_indices:
        worksheet.cell(row=1, column=col).fill = green_fill  # Применяем зеленый фон к указанным заголовкам

    for col in blue_header_indices:
        worksheet.cell(row=1, column=col).fill = blue_fill  # Применяем зеленый фон к указанным заголовкам

    for col in peach_puff_indices:
        worksheet.cell(row=1, column=col).fill = peach_puff_fill  # Применяем зеленый фон к указанным заголовкам

    for col in moccasin_fill_indices:
        worksheet.cell(row=1, column=col).fill = moccasin_fill  # Применяем зеленый фон к указанным заголовкам

    # Установка переноса текста для заголовков и жирного шрифта
    bold_font = Font(bold=True)  # Создаем объект шрифта с жирным начертанием

    # Создание границ
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    for col in range(1, len(headers) + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')  # Устанавливаем перенос текста
        cell.font = bold_font  # Устанавливаем жирный шрифт
        cell.border = thin_border  # Применяем границы к ячейке заголовка


        # Заполнение данными

    for item in data_smp:
        row_data = [
            item.data_vyzova_smp.strftime("%d.%m.%Y") if item.data_vyzova_smp else '', #
            item.fio_pacienta,                                                          #
            item.kolichestvo_vyezdov_v_den,                                             #
            item.data_rozhdeniya.strftime("%d.%m.%Y") if item.data_rozhdeniya else '', #
            item.polis_oms,
            item.rezultat_vyzova,
            item.aktiv_passiv,
            item.diagnoz_po_mkb,
            item.prichina_vyzova_skoroiy_pomoschi,
                        item.nalichie_bolevogo_sindroma,
            item.prichina_vyzova_skoroiy_pomoschi_kratko,
            item.kuriruyushchee_podrazdelenie_ovpp,
            item.data_vklyucheniya_v_registr.strftime("%d.%m.%Y") if item.data_vklyucheniya_v_registr else '',
            item.data_poslednego_vizita_vracha_iz_protokola_osmotra_emias.strftime(
                "%d.%m.%Y") if item.data_poslednego_vizita_vracha_iz_protokola_osmotra_emias else '',
            item.fio_vracha,
            item.sravnenie_povoda_vyz_smp_s_prot_vracha_cpp_do_i_posle,
            item.dejstviya_cpp,
            item.kontrol_ispolneniya_naznachennykh_vrachom_cpp_rekomendacij,
            item.vyvod,

            item.kakova_prichina_vyzova_smp_po_rezutatam_audiokontrolya,
            item.kakie_dejstviya_byli_predprinyaty_smp,
            item.sootvetstvuet_li_naznachennyj_bazovyj_plan_sostoyaniyu_pacie,
            item.trebovanie_gospitalizacii_na_dannyj_moment,
            item.vyyavlennye_defekty_v_rabote_vracha,
            item.bazovyj_plan_naznachen_korrektno,
            item.ocenka_sostoyaniya_sootvetstvuet_bazovomu_planu,
            item.zhaloby_opisany_v_polnom_obeme,
            item.ocenka_zaveduyushchego_proizvedena_korrektno,
            item.ocenka_dejstvij_vracha_do_vyzova_smp,
            item.ocenka_dejstvij_posle_vyzova_smp,
            item.vyvody_po_rezultatm_ocenki,

        ]

        # Добавляем данные в строку
        worksheet.append(row_data)

        # Применяем границы ко всем ячейкам в текущей строке
        current_row = worksheet.max_row  # Получаем номер последней строки
        for col in range(1, len(row_data) + 1):
            cell = worksheet.cell(row=current_row, column=col)
            cell.border = thin_border  # Применяем границы к ячейке

    # Сохранение файла в памяти
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    # ----- Формирую название файла --------
    if search_data_vizova_smp:
        file = f'smp_{search_data_vizova_smp}.xlsx'
    else:      file = 'smp_all.xlsx'
    print(file)

    # Создание HTTP ответа
    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{file}"'
    return response